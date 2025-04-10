import openpyxl
import os
import pandas as pd
import seaborn
from openpyxl.styles import Font, PatternFill, Alignment

#Definimos la función de crear excel
def crear_excel_inicial():
    #Creamos un nuevo libro de trabajo
    wb = openpyxl.Workbook()
    #Creamos una hoja de trabajo activa
    ws = wb.active
    ws.title = "Inventario principal"

    #Definimos los títulos de cada columna
    encabezados = ['ID', 'Nombre', 'Descripción', 'Cantidad', 'Precio Unitario', 'Precio Total']

    #Rellenamos con datos aleatorios
    datos = [
        [1, 'Producto A', 'Descripción A', 1, 5.0, 50.0],
        [2, 'Producto B', 'Descripción B', 20, 3.0, 60.0],
        [3, 'Producto C', 'Descripción C', 3, 4.0, 60.0],
        [4, 'Producto D', 'Descripción D', 30, 2.0, 60.0],
        [5, 'Producto E', 'Descripción E', 25, 1.0, 25.0]
    ]

    #Escribimos los encabezados en la primera fila
    for col_num, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num, value=encabezado)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    #Rellenamos los datos
    for row_num, fila in enumerate(datos, 2):
        for col_num, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col_num, value=valor)
            cell.alignment = Alignment(horizontal='center')

    #Ajustamos ancho de las filas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    #Guardamos el archivo
    wb.save('Inventario.xlsx')

if __name__ == '__main__':
    #Verificamos si el archivo existe
    if not os.path.exists('Inventario.xlsx'):
        crear_excel_inicial()
        print("Archivo creado")
    else:
        print("El archivo ya existe")