#Sección de gráficos
#Importar librerías
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart, PieChart, ScatterChart
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
import os
import win32com.client as win32

#Creamos función para gráfico de ventas
def crear_grafico_ventas(ws):
    """Crea un gráfico de barras para las ventas por mes."""
    #creamos un gráfico de barras
    grafico = BarChart()
    grafico.title = "Cantidad de productos"
    grafico.x_axis.title = "Productos"
    grafico.y_axis.title = "Cantidad de productos"
    grafico.style = 10

    #Datos de gráfico
    datos = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row, max_col=4)
    categorias = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(categorias)

    #Añadimos gráfico a la hoja
    ws.add_chart(grafico, "H3")


    #Se crea un segundo gráfico
def crear_grafico_categorias(ws):
    """Crea un gráfico de torta para validar la distribución de categorías"""
    #Creamos grafico de torta
    grafico = PieChart()
    grafico.title = "Distribución de categorías"
    grafico.style = 10

    #agregamos datos al gráfico
    datos = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row, max_col=3)
    categorias = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)


#Creamos función para llamar los gráficos
def automatizacion_avanzada():
    """Función principal para cargar la automatización"""
    
    #Cargamos el archivo existente
    wb = openpyxl.load_workbook("Inventario.xlsx")
    ws = wb.active
    print("---------------------------------")
    print("Cargando archivo...")
    print("---------------------------------")
    print("Creando gráficos automaticos")
    crear_grafico_ventas(ws)
    print("---------------------------------")
    #Guardamos cambios
    wb.save("InventarioV3.xlsx")

if __name__ == "__main__":
    automatizacion_avanzada()