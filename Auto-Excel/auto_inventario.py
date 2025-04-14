import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime

def cargar_inventario():
    try:
        # Cargar el archivo de Excel existente
        wb = openpyxl.load_workbook('Inventario.xlsx')
        #return wb, ws
    except FileNotFoundError:
        print("El archivo 'Inventario.xlsx' no existe. Creando uno nuevo.")
        return None
    
    ws = wb.active
    ws.title = 'Inventario principal'

    #Aplicamos formato a los encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='000000'),
            right=openpyxl.styles.Side(style='thin', color='000000'),
            top=openpyxl.styles.Side(style='thin', color='000000'),
            bottom=openpyxl.styles.Side(style='thin', color='000000')
        )
    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    return wb, ws



def actualizar_precios(ws, porcentaje_incremento):
    """Actualizar los precios unitarios y recalcular los precios totales"""
    for row in ws.iter_rows(min_row=2):  # Saltamos encabezado
        cantidad_cell = row[3]  # Columna D - Cantidad
        precio_cell = row[4]    # Columna E - Precio Unitario
        total_cell = row[5]     # Columna F - Precio Total

        try:
            cantidad = float(cantidad_cell.value) if cantidad_cell.value is not None else 0
            precio_actual = float(precio_cell.value)
            nuevo_precio = precio_actual * (1 + porcentaje_incremento / 100)
            # Actualizar precio unitario
            precio_cell.value = round(nuevo_precio, 2)
            # Recalcular y actualizar precio total
            total_cell.value = round(cantidad * nuevo_precio, 2)
        except ValueError:
            print(f"No se pudo procesar la fila {precio_cell.row}. Verifica que los datos sean numéricos.")


#Función para alertar si la cantidad es baja
def verificar_stock(ws, umbral=5):
    """Verificar si la cantidad de productos es menor que el umbral"""
    for row in ws.iter_rows(min_row=2):  # Saltamos encabezado
        cantidad_cell = row[3]  # Columna D - Cantidad
        try:
            cantidad = float(cantidad_cell.value) if cantidad_cell.value is not None else 0
            if cantidad < umbral:
                print(f"Alerta: El producto en la fila {cantidad_cell.row} tiene bajo stock ({cantidad}).")
                #colocamos el color rojo en esa celda
                cantidad_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        except ValueError:
            print(f"No se pudo procesar la fila {cantidad_cell.row}. Verifica que los datos sean numéricos.")


#Creamos un infome diario con la información del inventario
def generar_reporte_diario(ws):
    """Generar un informe diario del inventario"""
    wb, ws = cargar_inventario()

    if not wb:
        return
    #crear hoja nueva con reporte
    fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")
    reporte_ws = wb.create_sheet(title=f"Reporte {fecha_actual}")
    # Copiar encabezados
    for col_num, cell in enumerate(ws[1], 1):
        reporte_ws.cell(row=1, column=col_num, value=cell.value).font = Font(bold=True)
    # Copiar datos
    for row in ws.iter_rows(min_row=2, values_only=True):
        reporte_ws.append(row)

    
    #añadir estadísticas, esta sección es para crear estadísticas
    row_stats = ws.max_row + 2
    reporte_ws.cell(row=row_stats, column=1, value="Estadísticas")
    reporte_ws.cell(row=row_stats, column=1).font = Font(bold=True)
    reporte_ws.cell(row=row_stats, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    reporte_ws.cell(row=row_stats + 1, column=1, value="Productos Diferentes:")
    reporte_ws.cell(row=row_stats + 1, column=2, value=f"=COUNTA(B2:B{ws.max_row})")
    reporte_ws.cell(row=row_stats + 2, column=1, value="Total Productos:")
    reporte_ws.cell(row=row_stats + 2, column=2, value=f"=SUM(D2:B{ws.max_row})")
    reporte_ws.cell(row=row_stats + 3, column=1, value="Valor Total del Inventario:")
    reporte_ws.cell(row=row_stats + 3, column=2, value=f"=SUM(F2:F{ws.max_row})")
    
    return wb


def automatizacion_inventario():
    """Función para automatizar las operaciones que realiza el inventario"""

    wb, ws = cargar_inventario()
    if not wb:
        return
    print("Sistema de Automatización de inventario")

    #Acá se actualizan los precios de los productos
    actualizar_precios(ws, porcentaje_incremento=5) #Ejemplo de actualización de precios
    print("Actualizando precios de productos...")

    #Acá se verifica el stock de los productos
    verificar_stock(ws, umbral=5) #Ejemplo de verificación de stock
    print("Verificando stock de productos...")

    wb = generar_reporte_diario(datetime.datetime.now().strftime("%Y-%m-%d"))
    if not wb:
        return
    
    wb.save('Inventario.xlsx')
    print("El archivo 'Inventario.xlsx' ha sido creado y guardado exitosamente.")

if __name__ == "__main__":
    automatizacion_inventario()

